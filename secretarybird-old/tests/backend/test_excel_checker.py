"""Tests for the Excel/Sheets error checker module."""
import os
import tempfile
import pytest
from openpyxl import Workbook

from backend.core.excel_checker import ExcelChecker, ExcelError, ErrorSeverity


@pytest.fixture
def checker():
    """Create an ExcelChecker instance."""
    return ExcelChecker()


@pytest.fixture
def sample_workbook(tmp_path):
    """Create a sample Excel workbook with various errors for testing."""
    filepath = tmp_path / "test_errors.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Row 1: Headers
    ws["A1"] = "Name"
    ws["B1"] = "Amount"
    ws["C1"] = "Date"
    ws["D1"] = "Formula"

    # Row 2: Normal data
    ws["A2"] = "Alice"
    ws["B2"] = 100.50
    ws["C2"] = "2024-01-15"
    ws["D2"] = "=B2*2"

    # Row 3: Normal data
    ws["A3"] = "Bob"
    ws["B3"] = 200.75
    ws["C3"] = "2024-02-20"
    ws["D3"] = "=B3*2"

    # Row 4: Mixed type in Amount column (string instead of number)
    ws["A4"] = "Charlie"
    ws["B4"] = "not a number"
    ws["C4"] = "2024-03-10"
    ws["D4"] = "=B4*2"

    # Row 5: Empty cell in expected range
    ws["A5"] = "Diana"
    # B5 intentionally left empty
    ws["C5"] = "invalid-date"
    ws["D5"] = "=B5*2"

    # Row 6: Data with inconsistent formula
    ws["A6"] = "Eve"
    ws["B6"] = 500
    ws["C6"] = "2024-05-01"
    ws["D6"] = "=B6+100"  # Different formula pattern than D2-D5

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def error_formula_workbook(tmp_path):
    """Create workbook with Excel error values."""
    filepath = tmp_path / "test_formula_errors.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"

    # We can't directly write error values via openpyxl in the same way
    # Excel does, but we can test that the checker handles normal files
    ws["A1"] = "Value"
    ws["A2"] = 10
    ws["A3"] = 0
    ws["B1"] = "Result"
    ws["B2"] = "=A2/A3"  # Would be #DIV/0! in Excel
    ws["B3"] = "=A2/A2"

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def empty_cells_workbook(tmp_path):
    """Create workbook with empty cells in mostly-populated columns (>90% fill)."""
    filepath = tmp_path / "test_empty.xlsx"
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Score"

    # 11 data rows with 1 empty in B -> ~91% fill rate, triggers detection
    for i in range(2, 13):
        ws[f"A{i}"] = i - 1
        ws[f"C{i}"] = 90 + i
    for i in range(2, 13):
        ws[f"B{i}"] = f"Person{i}"
    # Make one cell empty (1 out of 11 = 9% empty, 91% fill)
    ws["B7"] = None

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def mixed_types_workbook(tmp_path):
    """Create workbook with mixed data types in columns."""
    filepath = tmp_path / "test_mixed.xlsx"
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ID"
    ws["B1"] = "Value"

    # 10 data rows: 7 numbers, 3 strings in B column -> 30% minority, triggers detection
    for i in range(2, 9):
        ws[f"A{i}"] = i - 1
        ws[f"B{i}"] = (i - 1) * 100
    ws["A9"] = 8
    ws["B9"] = "eight hundred"
    ws["A10"] = 9
    ws["B10"] = "nine hundred"
    ws["A11"] = 10
    ws["B11"] = "one thousand"

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def csv_file(tmp_path):
    """Create a sample CSV file with issues."""
    filepath = tmp_path / "test_data.csv"
    filepath.write_text(
        "Name,Amount,Date\n"
        "Alice,100.50,2024-01-15\n"
        "Bob,200.75,2024-02-20\n"
        "Charlie,not_a_number,2024-03-10\n"
        "Diana,,invalid-date\n"
        "Eve,500,2024-05-01\n"
    )
    return str(filepath)


class TestExcelChecker:
    """Test suite for ExcelChecker."""

    def test_init(self, checker):
        """Test checker initialization."""
        assert checker is not None

    def test_check_xlsx_file(self, checker, sample_workbook):
        """Test checking an xlsx file returns a report."""
        report = checker.check_file(sample_workbook)
        assert report is not None
        assert "file" in report
        assert "sheets" in report
        assert report["file"] == sample_workbook

    def test_detect_empty_cells(self, checker, empty_cells_workbook):
        """Test detection of empty cells in data ranges."""
        report = checker.check_file(empty_cells_workbook)
        warnings = report["sheets"][0]["warnings"]
        empty_warnings = [w for w in warnings if w["type"] == "empty_cells_in_range"]
        assert len(empty_warnings) > 0

    def test_detect_mixed_types(self, checker, mixed_types_workbook):
        """Test detection of mixed data types in columns."""
        report = checker.check_file(mixed_types_workbook)
        warnings = report["sheets"][0]["warnings"]
        mixed_warnings = [w for w in warnings if w["type"] == "mixed_data_types"]
        assert len(mixed_warnings) > 0

    def test_check_csv_file(self, checker, csv_file):
        """Test checking a CSV file."""
        report = checker.check_file(csv_file)
        assert report is not None
        assert "file" in report

    def test_csv_detect_empty_values(self, checker, tmp_path):
        """Test CSV empty value detection."""
        # Need >90% fill rate to trigger detection, so 11 rows with 1 empty
        filepath = tmp_path / "csv_empty.csv"
        lines = ["Name,Amount,Date\n"]
        for i in range(1, 12):
            lines.append(f"Person{i},{i*100},2024-0{min(i,9)}-{10+i}\n")
        # Make one Amount empty (1/11 = 9% empty, 91% fill)
        lines[5] = "Person5,,2024-05-15\n"
        filepath.write_text("".join(lines))
        report = checker.check_file(str(filepath))
        warnings = report["sheets"][0]["warnings"]
        empty_warnings = [w for w in warnings if w["type"] == "empty_cells_in_range"]
        assert len(empty_warnings) > 0

    def test_csv_detect_type_issues(self, checker, tmp_path):
        """Test CSV type inconsistency detection."""
        # Need enough rows for mixed type detection (>2% minority)
        filepath = tmp_path / "csv_mixed.csv"
        lines = ["Name,Amount\n"]
        for i in range(1, 9):
            lines.append(f"Person{i},{i*100}\n")
        # Add text values in Amount column (3 out of 10 = 30% minority)
        lines.append("Person9,not_a_number\n")
        lines.append("Person10,also_text\n")
        lines.append("Person11,more_text\n")
        filepath.write_text("".join(lines))
        report = checker.check_file(str(filepath))
        warnings = report["sheets"][0]["warnings"]
        mixed_warnings = [w for w in warnings if w["type"] == "mixed_data_types"]
        assert len(mixed_warnings) > 0

    def test_inconsistent_formulas(self, checker, sample_workbook):
        """Test detection of inconsistent formulas."""
        report = checker.check_file(sample_workbook)
        warnings = report["sheets"][0]["warnings"]
        formula_warnings = [w for w in warnings if w["type"] == "inconsistent_formula"]
        assert len(formula_warnings) > 0

    def test_nonexistent_file(self, checker):
        """Test handling of nonexistent file."""
        with pytest.raises(FileNotFoundError):
            checker.check_file("/nonexistent/file.xlsx")

    def test_unsupported_format(self, checker, tmp_path):
        """Test handling of unsupported file format."""
        bad_file = tmp_path / "test.txt"
        bad_file.write_text("not a spreadsheet")
        with pytest.raises(ValueError):
            checker.check_file(str(bad_file))

    def test_error_severity_levels(self):
        """Test that error severity enum has expected values."""
        assert ErrorSeverity.ERROR is not None
        assert ErrorSeverity.WARNING is not None
        assert ErrorSeverity.INFO is not None

    def test_report_summary(self, checker, sample_workbook):
        """Test that report includes a summary."""
        report = checker.check_file(sample_workbook)
        assert "summary" in report
        assert "total_errors" in report["summary"]
        assert "total_warnings" in report["summary"]


class TestExcelError:
    """Test the ExcelError dataclass."""

    def test_create_error(self):
        """Test creating an ExcelError."""
        error = ExcelError(
            type="empty_cell",
            severity=ErrorSeverity.WARNING,
            sheet="Sheet1",
            cell="B5",
            message="Empty cell in data range",
        )
        assert error.type == "empty_cell"
        assert error.severity == ErrorSeverity.WARNING
        assert error.sheet == "Sheet1"
        assert error.cell == "B5"

    def test_error_to_dict(self):
        """Test converting error to dictionary."""
        error = ExcelError(
            type="mixed_types",
            severity=ErrorSeverity.ERROR,
            sheet="Sheet1",
            cell="B4",
            message="Mixed data types in column B",
        )
        d = error.to_dict()
        assert isinstance(d, dict)
        assert d["type"] == "mixed_types"
        assert d["severity"] == "error"
