"""
Excel / Spreadsheet Error Checker

Detects common issues in Excel (.xlsx) and CSV files:
- Excel formula errors (#REF!, #VALUE!, #NAME?, #NULL!, #N/A, #DIV/0!, #NUM!)
- Circular references
- Inconsistent formulas within columns
- Mixed data types in columns
- Empty cells within expected data ranges
- Potential date format issues
- Returns structured error reports
"""

import csv
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Severity enum and ExcelError dataclass
# ---------------------------------------------------------------------------

class ErrorSeverity(str, Enum):
    """Severity levels for detected issues."""
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


@dataclass
class ExcelError:
    """Structured representation of a spreadsheet issue."""
    type: str
    severity: ErrorSeverity
    sheet: str = ""
    cell: str = ""
    message: str = ""

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["severity"] = self.severity.value
        return d


logger = logging.getLogger("tojo.excel_checker")

# ---------------------------------------------------------------------------
# Known Excel error values
# ---------------------------------------------------------------------------
EXCEL_ERROR_VALUES: set[str] = {
    "#REF!",
    "#VALUE!",
    "#NAME?",
    "#NULL!",
    "#N/A",
    "#DIV/0!",
    "#NUM!",
}

# Patterns that look like dates but might be ambiguous (e.g., 01/02/03)
_AMBIGUOUS_DATE_RE = re.compile(
    r"^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$"
)


class ExcelChecker:
    """
    Checks Excel (.xlsx) and CSV files for common data quality issues
    and returns structured reports.
    """

    def check_file(self, file_path: str) -> dict[str, Any]:
        """
        Run all checks on the given file.

        Args:
            file_path: Path to an .xlsx or .csv file.

        Returns:
            Structured report dictionary.
        """
        path = Path(file_path).resolve()
        if not path.is_file():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = path.suffix.lower()
        if ext in (".xlsx", ".xlsm", ".xlsb"):
            return self._check_excel(path)
        elif ext in (".csv", ".tsv"):
            return self._check_csv(path, delimiter="\t" if ext == ".tsv" else ",")
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    # ------------------------------------------------------------------
    # Excel (.xlsx) checks
    # ------------------------------------------------------------------
    def _check_excel(self, path: Path) -> dict[str, Any]:
        """Run all checks on an Excel workbook."""
        wb = openpyxl.load_workbook(str(path), data_only=False)
        wb_data = openpyxl.load_workbook(str(path), data_only=True)

        report: dict[str, Any] = {
            "file": str(path),
            "file_type": "excel",
            "sheets": [],
            "summary": {
                "total_errors": 0,
                "total_warnings": 0,
            },
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_data = wb_data[sheet_name]
            sheet_report = self._check_excel_sheet(ws, ws_data, sheet_name)
            report["sheets"].append(sheet_report)
            report["summary"]["total_errors"] += sheet_report["error_count"]
            report["summary"]["total_warnings"] += sheet_report["warning_count"]

        wb.close()
        wb_data.close()
        return report

    def _check_excel_sheet(
        self,
        ws: Any,
        ws_data: Any,
        sheet_name: str,
    ) -> dict[str, Any]:
        """Check a single Excel worksheet."""
        errors: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            return {
                "sheet": sheet_name,
                "rows": max_row,
                "columns": max_col,
                "errors": [],
                "warnings": [],
                "error_count": 0,
                "warning_count": 0,
            }

        # Collect column data for multi-cell checks
        col_formulas: dict[int, list[Optional[str]]] = defaultdict(list)
        col_types: dict[int, list[str]] = defaultdict(list)
        col_values: dict[int, list[tuple[int, Any]]] = defaultdict(list)

        # ------- Cell-level checks -------
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_data = ws_data.cell(row=row_idx, column=col_idx)
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

                value = cell.value
                data_value = cell_data.value

                # 1. Formula error values
                if isinstance(data_value, str) and data_value.strip() in EXCEL_ERROR_VALUES:
                    errors.append({
                        "type": "formula_error",
                        "cell": cell_ref,
                        "error_value": data_value.strip(),
                        "formula": str(value) if isinstance(value, str) and value.startswith("=") else None,
                        "severity": "error",
                    })

                # Track formulas for consistency checks (skip header row)
                if isinstance(value, str) and value.startswith("="):
                    col_formulas[col_idx].append(self._normalize_formula(value, row_idx))
                else:
                    col_formulas[col_idx].append(None)

                # Track types (skip header row 1)
                if row_idx > 1:
                    type_name = self._classify_type(data_value)
                    col_types[col_idx].append(type_name)
                    col_values[col_idx].append((row_idx, data_value))

        # ------- Circular references -------
        circ_refs = self._detect_circular_references(ws)
        for ref in circ_refs:
            errors.append({
                "type": "circular_reference",
                "cell": ref,
                "severity": "error",
            })

        # ------- Inconsistent formulas in columns -------
        for col_idx, formulas in col_formulas.items():
            non_null = [f for f in formulas if f is not None]
            if len(non_null) < 2:
                continue
            counter = Counter(non_null)
            if len(counter) > 1:
                most_common_formula, most_common_count = counter.most_common(1)[0]
                for row_offset, formula in enumerate(formulas):
                    if formula is not None and formula != most_common_formula:
                        cell_ref = f"{get_column_letter(col_idx)}{row_offset + 1}"
                        warnings.append({
                            "type": "inconsistent_formula",
                            "cell": cell_ref,
                            "column": get_column_letter(col_idx),
                            "expected_pattern": most_common_formula,
                            "actual_pattern": formula,
                            "severity": "warning",
                        })

        # ------- Mixed data types -------
        for col_idx, types in col_types.items():
            non_empty = [t for t in types if t != "empty"]
            if len(non_empty) < 2:
                continue
            type_counter = Counter(non_empty)
            if len(type_counter) > 1:
                dominant_type, dominant_count = type_counter.most_common(1)[0]
                minority_pct = (1 - dominant_count / len(non_empty)) * 100
                if minority_pct > 2:  # Only flag if > 2% are different
                    warnings.append({
                        "type": "mixed_data_types",
                        "column": get_column_letter(col_idx),
                        "type_distribution": dict(type_counter),
                        "dominant_type": dominant_type,
                        "minority_percentage": round(minority_pct, 1),
                        "severity": "warning",
                    })

        # ------- Empty cells in data ranges -------
        empty_cells = self._detect_empty_cells_in_range(col_values, max_row, max_col)
        for ec in empty_cells:
            warnings.append(ec)

        # ------- Date format issues -------
        date_issues = self._detect_date_issues(col_values, max_col)
        for di in date_issues:
            warnings.append(di)

        return {
            "sheet": sheet_name,
            "rows": max_row,
            "columns": max_col,
            "errors": errors,
            "warnings": warnings,
            "error_count": len(errors),
            "warning_count": len(warnings),
        }

    # ------------------------------------------------------------------
    # CSV checks
    # ------------------------------------------------------------------
    def _check_csv(self, path: Path, delimiter: str = ",") -> dict[str, Any]:
        """Run checks on a CSV file."""
        errors: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []

        col_types: dict[int, list[str]] = defaultdict(list)
        col_values: dict[int, list[tuple[int, Any]]] = defaultdict(list)
        row_lengths: list[int] = []
        total_rows = 0

        with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            header = next(reader, None)
            if header is None:
                return {
                    "file": str(path),
                    "file_type": "csv",
                    "sheets": [{
                        "sheet": "main",
                        "rows": 0,
                        "columns": 0,
                        "errors": [],
                        "warnings": [],
                        "error_count": 0,
                        "warning_count": 0,
                    }],
                    "summary": {"total_errors": 0, "total_warnings": 0},
                }

            expected_cols = len(header)
            for row_idx, row in enumerate(reader, start=2):
                total_rows += 1
                row_lengths.append(len(row))

                # Inconsistent column count
                if len(row) != expected_cols:
                    errors.append({
                        "type": "inconsistent_columns",
                        "row": row_idx,
                        "expected": expected_cols,
                        "actual": len(row),
                        "severity": "error",
                    })

                for col_idx, value in enumerate(row):
                    # Excel error values appearing in CSV
                    if value.strip() in EXCEL_ERROR_VALUES:
                        col_letter = get_column_letter(col_idx + 1) if col_idx < 16384 else str(col_idx)
                        errors.append({
                            "type": "formula_error",
                            "cell": f"{col_letter}{row_idx}",
                            "error_value": value.strip(),
                            "severity": "error",
                        })

                    type_name = self._classify_type(value)
                    col_types[col_idx].append(type_name)
                    col_values[col_idx].append((row_idx, value))

        # Mixed types
        for col_idx, types in col_types.items():
            non_empty = [t for t in types if t != "empty"]
            if len(non_empty) < 2:
                continue
            type_counter = Counter(non_empty)
            if len(type_counter) > 1:
                dominant_type, dominant_count = type_counter.most_common(1)[0]
                minority_pct = (1 - dominant_count / len(non_empty)) * 100
                if minority_pct > 2:
                    col_name = header[col_idx] if col_idx < len(header) else f"Column {col_idx + 1}"
                    warnings.append({
                        "type": "mixed_data_types",
                        "column": col_name,
                        "type_distribution": dict(type_counter),
                        "dominant_type": dominant_type,
                        "minority_percentage": round(minority_pct, 1),
                        "severity": "warning",
                    })

        # Empty cells in data
        max_col = expected_cols
        empty_cells = self._detect_empty_cells_in_range(col_values, total_rows + 1, max_col)
        for ec in empty_cells:
            warnings.append(ec)

        # Date issues
        date_issues = self._detect_date_issues(col_values, max_col)
        for di in date_issues:
            warnings.append(di)

        sheet_report = {
            "sheet": "main",
            "rows": total_rows + 1,  # including header
            "columns": expected_cols,
            "errors": errors,
            "warnings": warnings,
            "error_count": len(errors),
            "warning_count": len(warnings),
        }

        return {
            "file": str(path),
            "file_type": "csv",
            "sheets": [sheet_report],
            "summary": {
                "total_errors": len(errors),
                "total_warnings": len(warnings),
            },
        }

    # ------------------------------------------------------------------
    # Helper methods
    # ------------------------------------------------------------------
    @staticmethod
    def _normalize_formula(formula: str, row: int) -> str:
        """
        Normalize a formula by replacing the row number with a placeholder.
        This allows comparing formula patterns across rows in the same column.

        E.g., '=SUM(A2:B2)' with row=2 -> '=SUM(A{R}:B{R})'
        """
        return formula.replace(str(row), "{R}")

    @staticmethod
    def _classify_type(value: Any) -> str:
        """Classify the data type of a cell value."""
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return "empty"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, datetime):
            return "datetime"
        if isinstance(value, str):
            stripped = value.strip()
            # Check if it looks numeric
            try:
                float(stripped.replace(",", ""))
                return "number"
            except (ValueError, AttributeError):
                pass
            # Check if it looks like a date
            if _AMBIGUOUS_DATE_RE.match(stripped):
                return "date_string"
            return "text"
        return "other"

    @staticmethod
    def _detect_circular_references(ws: Any) -> list[str]:
        """
        Detect potential circular references in a worksheet.

        This performs a lightweight heuristic check: if a formula references
        its own cell address. Full transitive circular reference detection
        would require building a dependency graph.
        """
        circular: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    # Direct self-reference check
                    if cell_ref in cell.value:
                        circular.append(cell_ref)
        return circular

    @staticmethod
    def _detect_empty_cells_in_range(
        col_values: dict[int, list[tuple[int, Any]]],
        max_row: int,
        max_col: int,
    ) -> list[dict[str, Any]]:
        """
        Detect empty cells that appear within otherwise populated columns.
        Only flags columns where the majority of cells are populated.
        """
        issues: list[dict[str, Any]] = []

        for col_idx, values in col_values.items():
            non_empty = [(r, v) for r, v in values if v is not None and str(v).strip() != ""]
            empty = [(r, v) for r, v in values if v is None or str(v).strip() == ""]
            total = len(values)

            if total == 0:
                continue

            fill_rate = len(non_empty) / total
            # Only flag if column is mostly filled (>90%) but has gaps
            if fill_rate > 0.90 and len(empty) > 0 and len(empty) <= total * 0.10:
                col_letter = get_column_letter(col_idx + 1) if col_idx < 16384 else str(col_idx)
                empty_rows = [r for r, _ in empty]
                issues.append({
                    "type": "empty_cells_in_range",
                    "column": col_letter,
                    "empty_cell_count": len(empty),
                    "empty_rows": empty_rows[:20],  # cap for readability
                    "fill_rate": round(fill_rate * 100, 1),
                    "severity": "warning",
                })

        return issues

    @staticmethod
    def _detect_date_issues(
        col_values: dict[int, list[tuple[int, Any]]],
        max_col: int,
    ) -> list[dict[str, Any]]:
        """
        Detect potential date format ambiguity issues.

        Flags columns where date-like strings use inconsistent separators
        or could be interpreted in multiple ways (e.g., MM/DD vs DD/MM).
        """
        issues: list[dict[str, Any]] = []

        for col_idx, values in col_values.items():
            date_strings: list[tuple[int, str]] = []
            for row_idx, v in values:
                if isinstance(v, str) and _AMBIGUOUS_DATE_RE.match(v.strip()):
                    date_strings.append((row_idx, v.strip()))

            if len(date_strings) < 2:
                continue

            # Check separator consistency
            separators: set[str] = set()
            ambiguous_count = 0
            for _, ds in date_strings:
                for sep in "/-.":
                    if sep in ds:
                        separators.add(sep)
                parts = re.split(r"[/\-\.]", ds)
                if len(parts) >= 2:
                    try:
                        first, second = int(parts[0]), int(parts[1])
                        # If both parts could be month or day, it is ambiguous
                        if 1 <= first <= 12 and 1 <= second <= 12 and first != second:
                            ambiguous_count += 1
                    except ValueError:
                        pass

            col_letter = get_column_letter(col_idx + 1) if col_idx < 16384 else str(col_idx)

            if len(separators) > 1:
                issues.append({
                    "type": "inconsistent_date_separators",
                    "column": col_letter,
                    "separators_found": list(separators),
                    "date_count": len(date_strings),
                    "severity": "warning",
                })

            if ambiguous_count > 0:
                issues.append({
                    "type": "ambiguous_date_format",
                    "column": col_letter,
                    "ambiguous_count": ambiguous_count,
                    "total_dates": len(date_strings),
                    "examples": [ds for _, ds in date_strings[:5]],
                    "severity": "warning",
                })

        return issues
